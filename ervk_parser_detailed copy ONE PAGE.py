from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import pandas as pd
import time
import datetime
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import json
from webdriver_manager.chrome import ChromeDriverManager
import glob

# ============================================================================
# КОНФИГУРАЦИЯ БРАУЗЕРА
# ============================================================================
def setup_browser():
    """Настройка и запуск браузера."""
    options = Options()
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--start-maximized")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    return driver, WebDriverWait(driver, 15)

# ============================================================================
# КЛЮЧЕВЫЕ ФУНКЦИИ РАБОТЫ С КАРТОЧКАМИ
# ============================================================================
def find_cards():
    """Находит ВСЕ карточки на странице."""
    try:
        # Улучшенный поиск карточек - ищем по более специфичным селекторам
        cards = driver.find_elements(By.XPATH, 
            "//div[contains(@class, 'css-s85nh6') or contains(@class, 'MuiPaper-root') "
            "or contains(@class, 'object-card')]"
        )
        
        # Фильтруем только те карточки, которые содержат номер
        filtered_cards = []
        for card in cards:
            try:
                text = card.text[:50] if card.text else ""
                if '№' in text and ('значительный' in text.lower() or 'низкий' in text.lower() or 'средний' in text.lower()):
                    filtered_cards.append(card)
            except:
                continue
        
        print(f"   Найдено карточек: {len(filtered_cards)}")
        return filtered_cards
        
    except Exception as e:
        print(f"   Ошибка поиска карточек: {e}")
        return []

def expand_card_simple(card_element):
    """Раскрывает карточку ПРОСТЫМ и НАДЕЖНЫМ способом через JS."""
    try:
        # Прокручиваем к карточке
        driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", card_element)
        time.sleep(0.5)
        
        # ПРОСТОЙ JS КЛИК - как в работающем парсере
        js_click = """
        var card = arguments[0];
        
        // Пробуем кликнуть на все изображения внутри карточки
        var images = card.getElementsByTagName('img');
        for (var i = 0; i < images.length; i++) {
            try {
                images[i].click();
            } catch(e) {}
        }
        
        // Также кликаем на саму карточку
        card.click();
        
        // Двойной клик для надежности
        var evt = new MouseEvent('dblclick', {
            bubbles: true,
            cancelable: true,
            view: window
        });
        card.dispatchEvent(evt);
        
        return true;
        """
        
        # Выполняем JS клик
        driver.execute_script(js_click, card_element)
        
        # Ждем загрузки раскрытой информации
        time.sleep(2)
        
        # Проверяем, раскрылась ли карточка
        try:
            card_text = card_element.text
            if 'Адрес объекта контроля:' in card_text or 'ИНН:' in card_text or 'Контролируемые лица' in card_text:
                return True
        except:
            pass
        
        return False
        
    except Exception as e:
        print(f"      Ошибка при раскрытии: {e}")
        return False

def expand_all_cards():
    """Раскрывает ВСЕ карточки на странице ПЕРЕД парсингом."""
    cards = find_cards()
    
    if not cards:
        print("   ⚠ Карточки не найдены")
        return False
    
    print(f"   Раскрываю {len(cards)} карточек...")
    
    for i, card in enumerate(cards):
        try:
            # Прокручиваем
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", card)
            time.sleep(0.3)
            
            # Раскрываем
            expand_card_simple(card)
                
            # Пауза между карточками
            time.sleep(0.5)
            
        except Exception as e:
            print(f"      Ошибка карточки {i+1}: {e}")
    
    # Даем время на загрузку всех данных
    print("   Жду загрузки раскрытых данных...")
    time.sleep(3)
    
    return True

def parse_card_data(card_element):
    """Парсит данные из раскрытой карточки."""
    data = {
        'cosId': None,
        'ФИО': None,
        'Полное наименование контролируемого лица': None,
        'ИНН': None,
        'ОГРН': None,
        'ОГРНИП': None,
        'Адрес объекта контроля': None,
        'Категория риска': None,
        'Тип объекта': None,
        'Вид контроля': None,
        'Вид объекта контроля': None,
        'Подвид объекта контроля': None,
        'Время сбора': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'Статус': 'Собрано',
        'Номер страницы': None
    }
    
    try:
        # Получаем весь текст карточки
        card_text = card_element.text
        
        # 1. Извлекаем номер карточки (cosId)
        match = re.search(r'№\s*(\d+)', card_text)
        if match:
            data['cosId'] = match.group(1)
        
        # 2. Категория риска
        if 'значительный риск' in card_text.lower():
            data['Категория риска'] = 'значительный'
        elif 'низкий риск' in card_text.lower():
            data['Категория риска'] = 'низкий'
        elif 'средний риск' in card_text.lower():
            data['Категория риска'] = 'средний'
        elif 'высокий риск' in card_text.lower():
            data['Категория риска'] = 'высокий'
        
        # 3. Тип объекта (из заголовка)
        title_match = re.search(r'№\s*\d+\s*(.+?)(?:\s*версия\s*\d+)?$', card_text, re.MULTILINE)
        if title_match:
            data['Тип объекта'] = title_match.group(1).strip()
        
        # 4. ВИД КОНТРОЛЯ - используем регулярные выражения
        control_match = re.search(r'Вид контроля:\s*(.+)', card_text)
        if control_match:
            data['Вид контроля'] = control_match.group(1).strip()
        
        # 5. ВИД ОБЪЕКТА КОНТРОЛЯ
        object_type_match = re.search(r'Вид объекта контроля:\s*(.+)', card_text)
        if object_type_match:
            data['Вид объекта контроля'] = object_type_match.group(1).strip()
        
        # 6. ПОДВИД ОБЪЕКТА КОНТРОЛЯ
        subtype_match = re.search(r'Подвид объекта контроля:\s*(.+)', card_text)
        if subtype_match:
            data['Подвид объекта контроля'] = subtype_match.group(1).strip()
        
        # 7. АДРЕС
        address_match = re.search(r'Адрес объекта контроля:\s*(.+)', card_text)
        if address_match:
            data['Адрес объекта контроля'] = address_match.group(1).strip()
        
        # 8. КОНТРОЛИРУЕМЫЕ ЛИЦА - ОСНОВНАЯ ЧАСТЬ
        try:
            # Способ 1: Ищем элемент с классом css-kific6-wordBreak (как вы указали)
            fio_elements = card_element.find_elements(By.XPATH, 
                ".//p[contains(@class, 'css-kific6-wordBreak')]"
            )
            
            for elem in fio_elements:
                fio_text = elem.text.strip()
                if fio_text and len(fio_text) > 5 and ' ' in fio_text:
                    data['ФИО'] = fio_text
                    data['Полное наименование контролируемого лица'] = fio_text
                    break
            
            # Способ 2: Если не нашли по классу, ищем в тексте карточки
            if not data['ФИО']:
                lines = card_text.split('\n')
                for line in lines:
                    line = line.strip()
                    if (len(line) > 8 and ' ' in line and 
                        line[0].isupper() and 
                        not any(marker in line for marker in ['ИНН:', 'ОГРН:', 'Адрес:', 'Вид:', 'Тип:'])):
                        words = line.split()
                        if 2 <= len(words) <= 4:
                            data['ФИО'] = line
                            data['Полное наименование контролируемого лица'] = line
                            break
        
        except Exception as e:
            print(f"      Ошибка поиска ФИО: {e}")
        
        # 9. ИНН и ОГРН - ищем по паттернам
        inn_match = re.search(r'ИНН\s*[:：]?\s*(\d{10,12})', card_text)
        if inn_match:
            data['ИНН'] = inn_match.group(1)
        
        ogrn_match = re.search(r'ОГРН\s*[:：]?\s*(\d{13})', card_text)
        if ogrn_match:
            data['ОГРН'] = ogrn_match.group(1)
        
        ogrnip_match = re.search(r'ОГРНИП\s*[:：]?\s*(\d{15})', card_text)
        if ogrnip_match:
            data['ОГРНИП'] = ogrnip_match.group(1)
            if not data['ОГРН']:
                data['ОГРН'] = data['ОГРНИП']
        
        # 10. Если ИНН/ОГРН не найдены, пробуем поискать в любом месте текста
        if not data['ИНН']:
            all_numbers = re.findall(r'\b\d{10,12}\b', card_text)
            for num in all_numbers:
                if len(num) in [10, 12]:
                    data['ИНН'] = num
                    break
        
        if not data['ОГРН']:
            all_numbers = re.findall(r'\b\d{13,15}\b', card_text)
            for num in all_numbers:
                if len(num) in [13, 15]:
                    data['ОГРН'] = num
                    break
        
        # 11. Статус сбора
        if data['ФИО'] and data['ИНН']:
            data['Статус'] = '✓ Успешно'
        elif data['ФИО']:
            data['Статус'] = '⚠ Только ФИО'
        elif data['ИНН']:
            data['Статус'] = '⚠ Только ИНН'
        else:
            data['Статус'] = '✗ Данных нет'
        
        return data
        
    except Exception as e:
        print(f"      Ошибка парсинга: {e}")
        data['Статус'] = f'Ошибка: {str(e)[:30]}'
        return data

def process_page(page_num):
    """Обрабатывает одну страницу и возвращает данные."""
    print(f"\n{'='*60}")
    print(f"📄 СТРАНИЦА {page_num}")
    print(f"{'='*60}")
    
    page_data = []
    
    try:
        # 1. Раскрываем ВСЕ карточки на странице
        print("1. Раскрываю все карточки на странице...")
        if not expand_all_cards():
            print("   ⚠ Не удалось раскрыть карточки")
            return page_data
        
        # 2. Находим карточки после раскрытия
        print("2. Ищу раскрытые карточки...")
        cards = find_cards()
        
        if not cards:
            print("   ⚠ Карточки не найдены после раскрытия")
            return page_data
        
        print(f"   Найдено {len(cards)} карточек для парсинга")
        
        # 3. Парсим каждую карточку
        print("3. Парсим данные...")
        
        for i, card in enumerate(cards):
            try:
                # Парсим данные
                card_data = parse_card_data(card)
                card_data['Номер страницы'] = page_num
                
                # Добавляем в данные страницы
                page_data.append(card_data)
                
                # Выводим краткий результат
                if i < 5:  # Показываем только первые 5 для логов
                    status = "✓" if card_data['Статус'] == '✓ Успешно' else "⚠" if '⚠' in card_data['Статус'] else "✗"
                    print(f"   Карточка {i+1}: {status} {card_data.get('ФИО', 'нет ФИО')[:20]}... | ИНН: {card_data.get('ИНН', 'нет')}")
                
                # Пауза между карточками
                time.sleep(0.3)
                
            except Exception as e:
                print(f"   Ошибка обработки карточки {i+1}: {e}")
        
        # Показываем статистику по странице
        success_count = sum(1 for d in page_data if d['Статус'] == '✓ Успешно')
        print(f"\n📊 Статистика страницы {page_num}:")
        print(f"   Всего карточек: {len(page_data)}")
        print(f"   Успешно собрано: {success_count}")
        print(f"   С ФИО: {sum(1 for d in page_data if d.get('ФИО'))}")
        print(f"   С ИНН: {sum(1 for d in page_data if d.get('ИНН'))}")
        
        return page_data
        
    except Exception as e:
        print(f"⚠ Критическая ошибка на странице {page_num}: {e}")
        return page_data

def save_to_excel(data_list, filename):
    """Сохраняет данные в Excel файл с правильной структурой."""
    
    # Столбцы в нужном порядке
    columns = [
        'Номер страницы',
        'cosId',
        'Категория риска',
        'Тип объекта',
        'ФИО',
        'Полное наименование контролируемого лица',
        'ИНН',
        'ОГРН',
        'ОГРНИП',
        'Адрес объекта контроля',
        'Вид контроля',
        'Вид объекта контроля',
        'Подвид объекта контроля',
        'Время сбора',
        'Статус'
    ]
    
    try:
        # Создаем DataFrame
        df = pd.DataFrame(data_list)
        
        # Добавляем недостающие колонки
        for col in columns:
            if col not in df.columns:
                df[col] = None
        
        # Упорядочиваем колонки
        df = df[columns]
        
        # Сохраняем в Excel
        df.to_excel(filename, index=False)
        
        # Форматируем файл
        try:
            wb = load_workbook(filename)
            ws = wb.active
            
            # Ширина столбцов
            col_widths = {
                'A': 12, 'B': 15, 'C': 15, 'D': 40, 'E': 30, 'F': 40,
                'G': 15, 'H': 20, 'I': 20, 'J': 50, 'K': 40,
                'L': 50, 'M': 50, 'N': 20, 'O': 15
            }
            
            for col, width in col_widths.items():
                ws.column_dimensions[col].width = width
            
            # Заголовки жирным и цветом
            fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            font = Font(color="FFFFFF", bold=True)
            
            for cell in ws[1]:
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
            
            # Автоперенос текста для всех ячеек
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Цвет строк по статусу
            status_colors = {
                '✓ Успешно': 'C6EFCE',  # Светло-зеленый
                '⚠ Только ФИО': 'FFEB9C',  # Светло-желтый
                '⚠ Только ИНН': 'FFEB9C',
                '✗ Данных нет': 'FFC7CE',  # Светло-красный
            }
            
            for row in range(2, ws.max_row + 1):
                status = ws.cell(row=row, column=15).value
                if status in status_colors:
                    fill = PatternFill(start_color=status_colors[status], end_color=status_colors[status], fill_type="solid")
                    for col in range(1, 16):
                        ws.cell(row=row, column=col).fill = fill
            
            # Сохраняем форматирование
            wb.save(filename)
            wb.close()
            
        except Exception as e:
            print(f"    Ошибка форматирования: {e}")
        
        print(f"   💾 Сохранено {len(data_list)} записей в {filename}")
        return True
        
    except Exception as e:
        print(f"    Ошибка сохранения в Excel: {e}")
        return False

def save_page_data(page_data, page_num):
    """Сохраняет данные страницы во временный файл."""
    if not page_data:
        return None
    
    temp_filename = f'temp_page_{page_num:03d}.xlsx'
    if save_to_excel(page_data, temp_filename):
        return temp_filename
    return None

def merge_all_pages(output_filename, temp_files):
    """Объединяет все временные файлы в один итоговый."""
    all_data = []
    
    for temp_file in temp_files:
        try:
            df = pd.read_excel(temp_file)
            all_data.append(df)
            print(f"   Загружено {len(df)} записей из {temp_file}")
        except Exception as e:
            print(f"   Ошибка загрузки {temp_file}: {e}")
    
    if all_data:
        # Объединяем все данные
        combined_df = pd.concat(all_data, ignore_index=True)
        
        # Сохраняем итоговый файл
        if save_to_excel(combined_df.to_dict('records'), output_filename):
            print(f"\n✅ Итоговый файл создан: {output_filename}")
            print(f"📊 Всего записей: {len(combined_df)}")
            
            # Статистика
            total_pages = combined_df['Номер страницы'].nunique()
            success_count = sum(1 for s in combined_df['Статус'] if s == '✓ Успешно')
            
            print(f"📈 Статистика:")
            print(f"   Всего страниц: {total_pages}")
            print(f"   Всего записей: {len(combined_df)}")
            print(f"   Успешно собрано: {success_count} ({success_count/len(combined_df)*100:.1f}%)")
            print(f"   С ФИО: {combined_df['ФИО'].notna().sum()}")
            print(f"   С ИНН: {combined_df['ИНН'].notna().sum()}")
            
            return True
    
    return False

def cleanup_temp_files(temp_files):
    """Удаляет временные файлы."""
    for temp_file in temp_files:
        try:
            if os.path.exists(temp_file):
                os.remove(temp_file)
                # print(f"   Удален {temp_file}")  # Можно раскомментировать для отладки
        except:
            pass

def find_next_page_button():
    """Находит кнопку перехода на следующую страницу."""
    try:
        # Пробуем разные селекторы для кнопки "следующая страница"
        selectors = [
            "//button[@aria-label='Перейти на следующую страницу']",
            "//button[contains(@class, 'fp-MuiPaginationItem-previousNext') and not(contains(@class, 'Mui-disabled'))]",
            "//button[.//*[contains(text(), '›') or contains(@data-testid, 'NavigateNextIcon')]]",
            "//button[contains(@class, 'MuiPaginationItem-root') and not(contains(@class, 'Mui-disabled')) and .//svg]"
        ]
        
        for selector in selectors:
            try:
                buttons = driver.find_elements(By.XPATH, selector)
                for button in buttons:
                    # Проверяем, что кнопка активна
                    if button.is_displayed() and button.is_enabled():
                        return button
            except:
                continue
        
        return None
        
    except Exception as e:
        print(f"   Ошибка поиска кнопки следующей страницы: {e}")
        return None

def wait_for_page_load(timeout=10):
    """Ждет загрузки страницы."""
    try:
        # Ждем появления карточек
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'css-s85nh6')]"))
        )
        return True
    except:
        print("   ⚠ Таймаут загрузки страницы, но продолжаем...")
        return True  # Все равно продолжаем

def get_current_page_number():
    """Пытается определить текущий номер страницы."""
    try:
        # Ищем активную страницу в пагинаторе
        active_pages = driver.find_elements(By.XPATH, 
            "//button[contains(@class, 'Mui-selected') and contains(@class, 'MuiPaginationItem-page')]"
        )
        
        if active_pages:
            page_text = active_pages[0].text
            if page_text.isdigit():
                return int(page_text)
        
        # Альтернативный способ - ищем в URL
        current_url = driver.current_url
        if 'page=' in current_url:
            match = re.search(r'page=(\d+)', current_url)
            if match:
                return int(match.group(1))
        
        return 1  # По умолчанию
        
    except:
        return 1

# ============================================================================
# ОСНОВНОЙ КОД
# ============================================================================
print("=" * 70)
print("ПАРСЕР ЕРВК - ПОЛНАЯ ВЕРСИЯ С ПАГИНАЦИЕЙ")
print("ОБРАБАТЫВАЕТ ВСЕ СТРАНИЦЫ АВТОМАТИЧЕСКИ")
print("=" * 70)

# Создаем имя файла
timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
output_filename = f'ЕРВК_все_страницы_{timestamp}.xlsx'
temp_files_dir = 'temp_pages'
os.makedirs(temp_files_dir, exist_ok=True)

print(f"📁 Итоговый файл: {output_filename}")
print(f"📁 Временные файлы: {temp_files_dir}/")
print("\n" + "=" * 70)

all_data = []
temp_files = []
driver = None
max_pages = 1000  # Максимальное количество страниц для безопасности

try:
    # 1. Настройка браузера
    print("\n1. Запускаю браузер...")
    driver, wait = setup_browser()
    
    # 2. Переход на сайт
    print("2. Открываю сайт https://ervk.gov.ru/objects...")
    driver.get("https://ervk.gov.ru/objects")
    time.sleep(3)
    
    # 3. Ручная настройка
    print("\n" + "=" * 70)
    print("ШАГ 1: РУЧНАЯ НАСТРОЙКА ПОИСКА")
    print("=" * 70)
    print("ВАЖНО: НЕ закрывайте браузер!")
    print("1. Настройте фильтры (регион, вид контроля и т.д.)")
    print("2. Дождитесь загрузки результатов")
    print("3. Нажмите Enter в этом окне")
    print("\nПрограмма автоматически обработает ВСЕ страницы")
    print("=" * 70)
    
    input("\nНажмите Enter, когда готовы...")
    
    # 4. Начинаем сбор данных
    print("\n2. Начинаю сбор данных со всех страниц...")
    print("   🔍 Будет обработано до 1000 страниц автоматически")
    print("   📊 Каждая страница сохраняется отдельно")
    print("   ⏳ Процесс может занять длительное время")
    print("\n" + "=" * 70)
    
    time.sleep(2)
    
    # 5. ОСНОВНОЙ ЦИКЛ ПО СТРАНИЦАМ
    current_page = get_current_page_number()
    processed_pages = 0
    
    while processed_pages < max_pages:
        print(f"\n{'='*60}")
        print(f"🚀 НАЧИНАЮ ОБРАБОТКУ СТРАНИЦЫ {current_page}")
        print(f"{'='*60}")
        
        # Обрабатываем текущую страницу
        page_data = process_page(current_page)
        
        if page_data:
            # Сохраняем данные страницы во временный файл
            temp_filename = os.path.join(temp_files_dir, f'page_{current_page:03d}.xlsx')
            if save_to_excel(page_data, temp_filename):
                temp_files.append(temp_filename)
                processed_pages += 1
                
                # Показываем прогресс
                print(f"\n✅ Страница {current_page} обработана и сохранена")
                print(f"📁 Файл: {temp_filename}")
            else:
                print(f"⚠ Ошибка сохранения страницы {current_page}")
        else:
            print(f"⚠ Страница {current_page} не содержит данных")
        
        # Пытаемся найти кнопку следующей страницы
        print(f"\n🔍 Ищу следующую страницу после {current_page}...")
        
        # Сохраняем текущий URL для проверки
        current_url_before = driver.current_url
        
        # Ищем кнопку следующей страницы
        next_button = find_next_page_button()
        
        if next_button:
            try:
                print(f"   Найдена кнопка следующей страницы")
                
                # Прокручиваем к кнопке
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_button)
                time.sleep(0.5)
                
                # Кликаем через JS для надежности
                driver.execute_script("arguments[0].click();", next_button)
                
                # Ждем загрузки новой страницы
                print(f"   Жду загрузки страницы {current_page + 1}...")
                time.sleep(3)
                
                # Проверяем, что страница действительно загрузилась
                wait_for_page_load(10)
                
                # Проверяем, что URL изменился
                current_url_after = driver.current_url
                if current_url_before != current_url_after:
                    print(f"   ✅ Успешный переход на страницу {current_page + 1}")
                    current_page += 1
                else:
                    print(f"   ⚠ URL не изменился, возможно это последняя страница")
                    break
                
                # Небольшая пауза перед обработкой следующей страницы
                time.sleep(1)
                
            except Exception as e:
                print(f"   ⚠ Ошибка при переходе на следующую страницу: {e}")
                break
        else:
            print(f"   ✅ Кнопка следующей страницы не найдена - это последняя страница")
            break
        
        # Ограничение на количество обработанных страниц (для безопасности)
        if processed_pages >= max_pages:
            print(f"\n⚠ Достигнут лимит в {max_pages} страниц")
            break
    
    # 6. ОБЪЕДИНЕНИЕ ВСЕХ СТРАНИЦ
    print("\n" + "=" * 70)
    print("ОБЪЕДИНЕНИЕ ДАННЫХ СО ВСЕХ СТРАНИЦ")
    print("=" * 70)
    
    if temp_files:
        print(f"\n📦 Объединяю данные из {len(temp_files)} страниц...")
        
        # Объединяем все временные файлы
        if merge_all_pages(output_filename, temp_files):
            print(f"\n🎉 ПАРСИНГ УСПЕШНО ЗАВЕРШЕН!")
            
            # Показываем примеры данных
            try:
                df = pd.read_excel(output_filename)
                print(f"\n📋 ПРИМЕРЫ СОБРАННЫХ ДАННЫХ:")
                print("-" * 80)
                
                sample = df.head(3)
                for idx, row in sample.iterrows():
                    print(f"Запись {idx+1} (страница {row.get('Номер страницы', '?')}):")
                    print(f"  cosId: {row.get('cosId')}")
                    print(f"  ФИО: {row.get('ФИО', 'не найдено')}")
                    print(f"  ИНН: {row.get('ИНН', 'не найдено')}")
                    print(f"  Адрес: {str(row.get('Адрес объекта контроля', 'не найден'))[:50]}...")
                    print(f"  Статус: {row.get('Статус', '?')}")
                    print()
                
                print("-" * 80)
                
            except Exception as e:
                print(f"Ошибка при чтении итогового файла: {e}")
        
        else:
            print("⚠ Ошибка объединения данных")
    else:
        print("⚠ Нет данных для объединения")
    
    # 7. ОЧИСТКА ВРЕМЕННЫХ ФАЙЛОВ
    print(f"\n🧹 Очищаю временные файлы...")
    cleanup_temp_files(temp_files)
    
    # Удаляем временную директорию если она пуста
    try:
        if os.path.exists(temp_files_dir) and not os.listdir(temp_files_dir):
            os.rmdir(temp_files_dir)
    except:
        pass

except KeyboardInterrupt:
    print("\n\n⚠ ПАРСИНГ ПРЕРВАН ПОЛЬЗОВАТЕЛЕМ!")
    
    # Сохраняем то, что успели собрать
    if temp_files:
        print(f"\n💾 Сохраняю собранные данные...")
        emergency_filename = f'ЕРВК_прервано_{timestamp}.xlsx'
        if merge_all_pages(emergency_filename, temp_files):
            print(f"✅ Данные сохранены в {emergency_filename}")
    
    cleanup_temp_files(temp_files)
    
except Exception as e:
    print(f"\n\n⚠ КРИТИЧЕСКАЯ ОШИБКА: {e}")
    import traceback
    traceback.print_exc()
    
finally:
    print("\n" + "=" * 70)
    print("ЗАВЕРШЕНИЕ РАБОТЫ")
    print("=" * 70)
    
    print("\n📋 ИТОГОВАЯ СТАТИСТИКА:")
    print(f"   Обработано страниц: {processed_pages}")
    print(f"   Сохранено временных файлов: {len(temp_files)}")
    
    print("\n📁 СОЗДАННЫЕ ФАЙЛЫ:")
    if os.path.exists(output_filename):
        print(f"   📄 {output_filename} - итоговый файл со всеми данными")
    
    print("\n🔧 РЕКОМЕНДАЦИИ:")
    print("1. Проверьте итоговый Excel файл")
    print("2. Если нужно продолжить с прерванного места:")
    print("   - Запустите парсер снова")
    print("   - Настройте фильтры на нужной странице")
    print("   - Программа продолжит с текущей страницы")
    print("3. Для больших объемов данных увеличьте timeout в wait_for_page_load()")
    print("=" * 70)
    
    if driver:
        input("\nНажмите Enter для закрытия браузера...")
        driver.quit()